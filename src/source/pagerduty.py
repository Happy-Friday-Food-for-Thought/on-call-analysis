import click
import requests

from config import configure_key
import datetime
from openpyxl import Workbook, load_workbook
import json
import pandas


def get_analytics(team_name, token, start_date, end_date=None):
    team_id = get_team_id(team_name, token)

    three_months_back = datetime.datetime.now() - datetime.timedelta(weeks=12)
    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d") if start_date else three_months_back
    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d") if end_date else datetime.datetime.now()

    url = 'https://api.pagerduty.com/analytics/metrics/incidents/teams'
    headers = {
        'Accept': 'application/vnd.pagerduty+json;version=2',
        'Content-Type': 'application/json',
        'Authorization': f'Token token={token}',
        'X-EARLY-ACCESS': 'analytics-v2'
    }

    body = {
        'filters': {
            'created_at_start': start_date.isoformat(),
            'created_at_end': end_date.isoformat(),
            'urgency': 'high',
            #  'major': True,
            'team_ids': [team_id]
        },
        'aggregate_unit': 'week',
        'time_zone': 'Etc/UTC'
    }

    response = requests.post(url, headers=headers, json=body)

    if response.status_code != 200:
        click.echo(f"Error fetching team aggregated data: {response.status_code}")

    return response.json()['data']


def get_incidents_log(team_name, token, start_date_str, end_date_str=None, incidents_limit=100, offset=0):
    team_id = get_team_id(team_name, token)

    three_months_back = datetime.datetime.now() - datetime.timedelta(weeks=12)
    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d") if start_date_str else three_months_back
    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d") if end_date_str else datetime.datetime.now()

    url = 'https://api.pagerduty.com/incidents'
    headers = {
        'Accept': 'application/vnd.pagerduty+json;version=2',
        'Authorization': f'Token token={token}'
    }

    params = {
        'since': start_date.isoformat(),
        'until': end_date.isoformat(),
        'team_ids[]': team_id,
        'offset': offset,
        'total': True,
        'limit': incidents_limit
    }

    response = requests.get(url, headers=headers, params=params)

    click.echo(f"Fetching {team_name} offset={offset}")
    if response.status_code != 200:
        click.echo(f"Error fetching team aggregated data: {response.status_code}")

    data = response.json()

    if not data['more']:
        return data['incidents']

    return data['incidents'] + get_incidents_log(team_name, token, start_date_str, end_date_str, incidents_limit, data['offset'] + incidents_limit)


def write_xls(stats, name):
    team_stats = [['Week Start Date', 'Incidents Count', 'Business-hour Interruptions', 'Off-hour Interruptions',
                   'Sleep-hour Interruptions', 'Engaged hours', 'Snoozed hours']]

    # Skip the first row (headers)
    for data in stats[1:]:
        week_start_string = data['range_start']
        week_start_date = date = datetime.datetime.strptime(week_start_string, "%Y-%m-%dT%H:%M:%S")
        row = [
            week_start_date.strftime('%Y-%m-%d'),
            data['total_incident_count'],
            data['total_business_hour_interruptions'],
            data['total_off_hour_interruptions'],
            data['total_sleep_hour_interruptions'],
            round(data['total_engaged_seconds'] / 3600, 2),
            round(data['total_snoozed_seconds'] / 3600, 2),
        ]
        team_stats.append(row)

    excel_file = f'./{name}'
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

    for row in team_stats:
        sheet.append(row)

    workbook.save(excel_file)


def get_team_id(team_name, pd_api_key):
    headers = {
        'Accept': 'application/vnd.pagerduty+json;version=2',
        'Content-Type': 'application/json',
        'Authorization': f'Token token={pd_api_key}',
        'X-EARLY-ACCESS': 'analytics-v2'
    }

    team_url = 'https://api.pagerduty.com/teams'
    team_params = {
        'query': team_name
    }
    team_response = requests.get(team_url, headers=headers, params=team_params)

    if team_response.status_code != 200:
        print(f"Error: {team_response.status_code} - {team_response.text}")
        return None

    teams = team_response.json()['teams']
    for team in teams:
        if team['name'].lower() == team_name.lower():
            return team['id']

    return None


@click.group()
def pagerduty():
    pass


@click.command(help="Get analytics data in the data/analytics.csv file for further analysis")
@click.option("--start-date", help="Format YYYY-mm-dd", default="")
@click.option("--end-date", help="Format YYYY-mm-dd", default="")
@click.option("--out", help="Path where to output the data", default="data/analytics.csv")
def analytics(start_date, end_date, out):
    token = configure_key("pagerduty.token", hide=True)
    team_name = configure_key("pagerduty.team")

    analytics_data = get_analytics(team_name, token, start_date, end_date)
    df = pandas.DataFrame(analytics_data)
    click.echo(df.describe())

    df.to_csv(path_or_buf=out)

    click.echo(f"Saved to {out}")


@click.command(help="Get incident log data in the data/incident_log.json file for further analysis")
@click.option("--start-date", help="Format YYYY-mm-dd", default="")
@click.option("--end-date", help="Format YYYY-mm-dd", default="")
@click.option("--out", help="Path where to output the data", default="data/incident_log.json")
def incidents_log(start_date, end_date, out):
    token = configure_key("pagerduty.token", hide=True)
    team_name = configure_key("pagerduty.team")

    incidents = get_incidents_log(team_name, token, start_date, end_date)
    with open(out, "w") as f:
        json.dump(incidents, f)

    click.echo(f"Saved to data/incident_log.json")


pagerduty.add_command(analytics)
pagerduty.add_command(incidents_log)
